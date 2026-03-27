"""
Logged Processing Wrappers
============================
Wrappers around the main engines that capture agent activity logs.
Returns results + activity log for the frontend agent panel.
"""

from agent_logger import new_logger
from reconciliation_sabseg import run_sabseg_reconciliation, parse_saldos_contables, ELEVIA_COMPANY_MAP
from data_quality import run_data_quality, validate_file
from corrections import apply_corrections, generate_broker_report, generate_all_reports
from qa_agents import QAOrchestrator
import pandas as pd
from io import BytesIO
import openpyxl
import os


def run_reconciliation_logged(saldos_bytes, stats_files, months=[1, 2]):
    """
    Run reconciliation with full agent activity logging.
    Returns: { result: {...}, agent_log: [...], qa_report: {...} }
    """
    log = new_logger()
    log.start('Reconciliación Contable — Cierre Ene-Feb 2026')
    
    # ── Agent: Ingestor — Read saldos ────────────────────────────────────
    log.log('ingestor', 'leyendo', f'Leyendo Saldos Contables (Business Central)...')
    
    try:
        wb = openpyxl.load_workbook(BytesIO(saldos_bytes), read_only=True)
        sheets = wb.sheetnames
        total_rows = 0
        empresas_preview = set()
        for sn in ['Ene-26', 'Feb-26']:
            if sn in sheets:
                ws = wb[sn]
                total_rows += ws.max_row or 0
                for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
                    if row[0]:
                        empresas_preview.add(str(row[0]).strip()[:40])
        wb.close()
        
        log.log('ingestor', 'completado', 
                f'Saldos Contables: {total_rows:,} filas, pestañas {[s for s in sheets if "26" in s]}',
                data={
                    'filas': total_rows,
                    'pestañas': [s for s in sheets if '26' in s],
                    'empresas_muestra': list(empresas_preview)[:5],
                })
    except Exception as e:
        log.log('ingestor', 'error', f'Error leyendo saldos: {str(e)}', icon='🔴')
    
    # ── Agent: Ingestor — Read statistics files ──────────────────────────
    log.log('ingestor', 'leyendo', f'Procesando {len(stats_files)} ficheros de estadísticas...')
    
    total_recibos = 0
    for filename, file_bytes in stats_files:
        size_kb = len(file_bytes) // 1024
        fn_upper = filename.upper()
        
        if fn_upper.endswith('.PDF'):
            log.log('ingestor', 'leyendo', f'📄 {filename} ({size_kb}KB) — Extrayendo texto de PDF...',
                    data={'filename': filename, 'size_kb': size_kb, 'type': 'PDF'})
        else:
            try:
                df_temp = pd.read_excel(BytesIO(file_bytes), sheet_name=0, nrows=0)
                ncols = len(df_temp.columns)
                wb_temp = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
                nrows = wb_temp[wb_temp.sheetnames[0]].max_row or 0
                wb_temp.close()
                total_recibos += nrows
                
                # Detect company
                empresa = 'desconocida'
                if 'ELEVIA' in fn_upper:
                    empresa = f'{len(ELEVIA_COMPANY_MAP)} empresas (ELEVIA)'
                else:
                    for key, val in {
                        'AGRINALCAZAR': 'Agrinalcázar', 'ARAYTOR': 'Araytor', 'ZURRIOLA': 'Zurriola',
                        'FUTURA': 'Futura', 'INSURART': 'Insurart', 'SANCHEZ': 'Sánchez Valencia',
                        'SEGURETXE': 'Seguretxe', 'VEROBROKER': 'Verobroker', 'ARRENTA': 'Arrenta',
                        'ADSA': 'ADSA Agro', 'AISA': 'AISA Agro', 'BANA': 'BANA Agro',
                    }.items():
                        if key in fn_upper:
                            empresa = val
                            break
                
                log.log('ingestor', 'leyendo', 
                        f'📊 {filename} → {nrows:,} filas, {ncols} columnas ({empresa})',
                        data={'filename': filename, 'filas': nrows, 'columnas': ncols, 'empresa': empresa})
            except:
                log.log('ingestor', 'leyendo', f'📊 {filename} ({size_kb}KB)')
    
    log.log('ingestor', 'completado', 
            f'Total: {len(stats_files)} ficheros ingestados, ~{total_recibos:,} recibos',
            data={'total_ficheros': len(stats_files), 'total_recibos': total_recibos}, icon='🟢')
    
    # ── Agent: Detector — Map companies ──────────────────────────────────
    log.log('detector', 'mapeando', 'Mapeando empresas de estadísticas a contabilidad...')
    
    mapped_empresas = []
    for code, name in sorted(ELEVIA_COMPANY_MAP.items()):
        mapped_empresas.append(f'{code} → {name[:35]}')
    
    log.log('detector', 'completado', 
            f'{len(ELEVIA_COMPANY_MAP)} empresas mapeadas de ELEVIA + {len(stats_files)-1} ficheros individuales',
            data={'mapeos': mapped_empresas[:8], 'total': len(ELEVIA_COMPANY_MAP)})
    
    # ── Run actual reconciliation ────────────────────────────────────────
    log.log('matching', 'cruzando', 'Cruzando comisiones estadísticas vs cuentas 705xxx y 623xxx...')
    
    result = run_sabseg_reconciliation(saldos_bytes, stats_files, months)
    
    if 'error' in result:
        log.log('matching', 'error', result['error'], icon='🔴')
        return {'result': result, 'agent_log': log.get_entries(), 'qa_report': None}
    
    # ── Agent: Matching — Log key results ────────────────────────────────
    recon = result.get('reconciliacion', [])
    matches = [r for r in recon if r['status'] == 'match']
    warnings = [r for r in recon if r['status'] == 'warning']
    mismatches = [r for r in recon if r['status'] == 'mismatch']
    
    # Show some example matches
    sample_results = []
    for r in recon[:6]:
        icon = '✓' if r['status'] == 'match' else ('⚠' if r['status'] == 'warning' else '✗')
        sample_results.append(
            f"{icon} {r['empresa'][:30]} {r['mes_label']} {r['cuenta']}: "
            f"Est {r['estadistica']:,.0f}€ vs Cont {r['contabilidad']:,.0f}€ → Δ {r['diferencia']:,.0f}€"
        )
    
    log.log('matching', 'completado',
            f'Cruce completado: {len(recon)} partidas analizadas',
            data={'muestra': sample_results})
    
    # ── Agent: Clasificador ──────────────────────────────────────────────
    log.log('clasificador', 'clasificando', 'Clasificando discrepancias por tipo y severidad...')
    
    log.log('clasificador', 'completado',
            f'{len(matches)} cuadradas ✓ | {len(warnings)} diferencia menor ⚠ | {len(mismatches)} discrepancias ✗',
            data={
                'cuadradas': len(matches),
                'warnings': len(warnings),
                'discrepancias': len(mismatches),
                'empresas': result.get('empresas_analizadas', 0),
            })
    
    # ── Agent: QA ────────────────────────────────────────────────────────
    log.log('qa', 'verificando', 'Verificando integridad de datos...', icon='🟡')
    
    qa = QAOrchestrator()
    qa_report = qa.run_reconciliation_qa(result)
    
    for agent_result in qa_report['agent_results']:
        passed = agent_result['checks_passed']
        total = agent_result['checks_run']
        fails = [c for c in agent_result['checks'] if c['status'] != 'pass']
        
        if fails:
            for f in fails:
                log.log('qa', 'check', f"[{f['status'].upper()}] {f['check']}: {f['detail'][:80]}", icon='🟡')
        else:
            log.log('qa', 'check', f"{agent_result['agent']}: {passed}/{total} checks ✓", icon='🟡')
    
    log.log('qa', 'completado',
            f"QA: {qa_report['qa_status'].upper()} — Confianza {qa_report['qa_confidence']} ({qa_report['pct_passed']}%)",
            data={'status': qa_report['qa_status'], 'confidence': qa_report['qa_confidence'], 'score': qa_report['pct_passed']},
            icon='🟡')
    
    # ── Finish ───────────────────────────────────────────────────────────
    summary = (f"Reconciliación completada. {result.get('empresas_analizadas', 0)} empresas, "
               f"{result.get('total_partidas', 0)} partidas. "
               f"Confianza QA: {qa_report['qa_confidence']} ({qa_report['pct_passed']}%)")
    
    agent_log = log.finish(summary)
    
    result['agent_log'] = agent_log
    result['qa_report'] = qa_report
    
    return result


def run_data_quality_logged(files):
    """
    Run data quality validation with full agent activity logging.
    """
    log = new_logger()
    log.start('Validación Data Quality — Ficheros de recibos')
    
    # ── Agent: Ingestor ──────────────────────────────────────────────────
    log.log('ingestor', 'leyendo', f'Procesando {len(files)} ficheros de recibos...')
    
    for filename, file_bytes in files:
        size_kb = len(file_bytes) // 1024
        try:
            wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
            sheets = wb.sheetnames
            main_sheet = sheets[0]
            ws = wb[main_sheet]
            nrows = ws.max_row or 0
            wb.close()
            
            correduria = 'desconocida'
            for key in ['ARAYTOR', 'ZURRIOLA', 'SEGURETXE', 'ARRENTA']:
                if key in filename.upper():
                    correduria = key.title()
                    break
            
            log.log('ingestor', 'leyendo',
                    f'📊 {filename} → {nrows:,} filas, {len(sheets)} pestañas ({correduria})',
                    data={'filename': filename, 'filas': nrows, 'correduria': correduria})
        except:
            log.log('ingestor', 'leyendo', f'📊 {filename} ({size_kb}KB)')
    
    log.log('ingestor', 'completado', f'{len(files)} ficheros ingestados', icon='🟢')
    
    # ── Agent: Validador estructura ──────────────────────────────────────
    log.log('validador_estructura', 'validando', 'Verificando estructura de ficheros (columnas, headers, formato)...')
    
    # ── Agent: Validador datos ───────────────────────────────────────────
    log.log('validador_datos', 'validando', 'Aplicando 20 reglas de validación del catálogo de errores...')
    
    # Run actual validation
    result = run_data_quality(files)
    
    # Log results per file
    for file_result in result.get('results', []):
        if 'error' in file_result:
            log.log('validador_datos', 'error', f"Error en {file_result.get('filename', '?')}: {file_result['error']}", icon='🔴')
            continue
        
        fn = file_result.get('filename', '?')
        corr = file_result.get('correduria', '?')
        n_err = file_result.get('total_errors', 0)
        n_warn = file_result.get('total_warnings', 0)
        n_auto = file_result.get('auto_correctable', 0)
        records = file_result.get('total_records', 0)
        
        # Log specific errors found
        error_summary = []
        for e in file_result.get('errors', []):
            error_summary.append(f"E{e.get('error_num', '?'):02d} {e['tipo']}: {e['cantidad']}")
        for w in file_result.get('warnings', []):
            error_summary.append(f"W{w.get('error_num', '?'):02d} {w['tipo']}: {w['cantidad']}")
        
        log.log('validador_datos', 'resultado',
                f'{corr}: {records:,} registros → {n_err} errores, {n_warn} warnings, {n_auto} auto-corregibles',
                data={'correduria': corr, 'filename': fn, 'errores': error_summary})
    
    log.log('validador_datos', 'completado',
            f"Total: {result.get('total_errors', 0)} errores, {result.get('total_warnings', 0)} warnings, {result.get('auto_correctable', 0)} auto-corregibles",
            icon='🟢')
    
    # ── Agent: QA ────────────────────────────────────────────────────────
    log.log('qa', 'verificando', 'Verificando calidad del análisis...', icon='🟡')
    
    qa = QAOrchestrator()
    qa_report = qa.run_data_quality_qa(result)
    
    for agent_result in qa_report['agent_results']:
        fails = [c for c in agent_result['checks'] if c['status'] != 'pass']
        if fails:
            for f in fails:
                log.log('qa', 'check', f"[{f['status'].upper()}] {f['detail'][:80]}", icon='🟡')
        else:
            log.log('qa', 'check', f"{agent_result['agent']}: {agent_result['checks_passed']}/{agent_result['checks_run']} checks ✓", icon='🟡')
    
    log.log('qa', 'completado',
            f"QA: {qa_report['qa_status'].upper()} — Confianza {qa_report['qa_confidence']} ({qa_report['pct_passed']}%)",
            icon='🟡')
    
    # ── Arrenta comparison ───────────────────────────────────────────────
    if result.get('arrenta_comparison'):
        ac = result['arrenta_comparison']
        log.log('validador_datos', 'comparacion',
                f"Arrenta Ene vs Feb: {ac.get('records_jan', 0)} → {ac.get('records_feb', 0)} registros",
                data={'enero': ac.get('records_jan', 0), 'febrero': ac.get('records_feb', 0), 'issues': ac.get('issues', [])})
    
    # ── Finish ───────────────────────────────────────────────────────────
    summary = (f"Validación completada. {result.get('total_files', 0)} ficheros, "
               f"{result.get('total_records', 0):,} registros. "
               f"Confianza QA: {qa_report['qa_confidence']} ({qa_report['pct_passed']}%)")
    
    agent_log = log.finish(summary)
    
    result['agent_log'] = agent_log
    result['qa_report'] = qa_report
    
    return result
